import tkinter as tk
from tkinter import filedialog
import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import img2pdf
import datetime

# Función para mostrar la alerta


# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("Automatización de PQR")
root.geometry("300x400")


# Texto de advertencia
warning_text = "Antes de ejecutar el programa, tenga en cuenta que este software ha sido diseñado exclusivamente para la radicación de PQR en IDIGER. Asegúrese de que el archivo Excel contenga únicamente 2 columnas con los títulos ASSET y RUTA ADJUNTO. No olvide cargar el driver de Selenium 'chromedriver'.\n\nSoftware diseñado por Luis Mellizo."

# Crear el marco para contener la etiqueta de advertencia
warning_frame = tk.Frame(root, bg="dark green")
warning_frame.pack(side="bottom", pady=10)

# Crear la etiqueta de advertencia
warning_label = tk.Label(warning_frame, text=warning_text,
                         justify="left", bg="dark green", fg="white", wraplength=280)
warning_label.pack()


# Estilo de fondo y colores
root.configure(bg="dark green")
button_bg = "orange"

ventana_adjunto = None


def abrir_excel():
    ruta_excel = filedialog.askopenfilename(
        filetypes=[("Archivos Excel", "*.xlsx")])
    # Asigna el valor a la variable excel_text y actualiza el campo de texto
    excel_text.set(ruta_excel)


def dime_donde_guardo():
    ruta_comprobante = filedialog.askdirectory()
    # Asigna el valor a la variable comprobante_text y actualiza el campo de texto
    comprobante_text.set(ruta_comprobante)


def obtener_ruta_chromedriver():
    ruta_chromedriver = filedialog.askopenfilename(
        filetypes=[("Chromedriver", "chromedriver.exe")])
    # Asigna el valor a la variable chromedriver_text y actualiza el campo de texto
    chromedriver_text.set(ruta_chromedriver)


def cargar_adjunto(driver, ruta_adjunto):
    # Obtener el elemento que contiene el botón "Examinar" por su ID
    div_archivos = driver.find_element(By.ID, 'divarchivos')

    # Hacer clic en el elemento que contiene el botón "Examinar"
    div_archivos.click()

    # Cargar el archivo adjunto
    file_adjunto_input = driver.find_element(By.ID, 'file1')
    file_adjunto_input.send_keys(ruta_adjunto)


def ejecutar_automatizacion():
    # Código de automatización
    # Obtiene la ruta del chromedriver desde el campo de texto
    driver = webdriver.Chrome(chromedriver_text.get())

    # Abre la página web
    driver.get('https://app1.sire.gov.co/Pqrs/')

    time.sleep(3)

    # Selección del tipo de solicitante
    tipo_solicitante_select = Select(driver.find_element(By.ID, 'tiposoli'))
    tipo_solicitante_select.select_by_value('52')  # Perona Natural

    # Localizar el elemento de texto "adulteraciones" y hacer clic en él
    contenedor_texto = driver.find_element(By.ID, 'tratadatos')
    driver.execute_script(
        "arguments[0].scrollTop = arguments[0].scrollHeight", contenedor_texto)
    time.sleep(0.3)

    # Seleccionar la casilla de aceptación
    aceptar_checkbox = driver.find_element(By.ID, 'checkbacepto')
    aceptar_checkbox.click()

    # Hacer clic en el botón "Siguiente"
    siguiente_button = driver.find_element(By.ID, 'enviarFormIndex')
    siguiente_button.click()

    # Espera hasta que aparezca el cuadro de diálogo
    wait = WebDriverWait(driver, 10)
    modal = wait.until(EC.visibility_of_element_located(
        (By.CSS_SELECTOR, "div.modal-content")))

    # Selecciona la opción "Sí"
    opcion_si = modal.find_element(By.CSS_SELECTOR, "input#tratadatossi")
    driver.execute_script("arguments[0].click();", opcion_si)

    time.sleep(0.3)

    # Haz clic en el botón "Aceptar"
    aceptar_button = driver.find_element(
        By.XPATH, "//input[@value='Aceptar' and contains(@class, 'btn btn-primary')]")
    aceptar_button.click()

    # Selección del tipo de documento
    tipo_documento_select = Select(
        driver.find_element(By.ID, 'select_t_document'))
    tipo_documento_select.select_by_value('40')  # Tarjeta de identidad

    # Introducir el número de cédula
    numero_cedula_input = driver.find_element(By.ID, 'numero_doc')
    numero_cedula_input.send_keys('1013689567')

    # Esperar 1 segundo para que se cargue la información del usuario
    time.sleep(1)

    # Ingresar el correo de confirmación
    correo_confirmacion_input = driver.find_element(By.ID, 'correoconfir')
    correo_confirmacion_input.send_keys('curaduria@serviciosespeciales.com.co')

    # Selección del tipo de solicitud
    tipo_solicitud_select = Select(driver.find_element(By.ID, 'tipo_soli'))
    tipo_solicitud_select.select_by_value('45')  # Petición

    # Selección del tema de solicitud
    tema_solicitud_select = Select(driver.find_element(By.ID, 'tema_soli'))
    tema_solicitud_select.select_by_value('20')  # Certificado riesgo y amenaza

    # Ingresar el asunto de la solicitud
    asunto_solicitud_input = driver.find_element(By.ID, 'asunto_sol')
    asunto_solicitud_input.send_keys('Se solicita al INSTITUTO DISTRITAL DE GESTION DE RIESGO Y CAMBIO CLIMATICO IDIGER, concepto en el que se indique si el predio en mención presenta afectación por inundación o deslizamiento y si esta presenta restricciones para la instalación, localización y regularización de estaciones radioeléctricas.')

    # Selección del modo de respuesta
    modo_respuesta_select = Select(driver.find_element(By.ID, 'modoRespuesta'))
    modo_respuesta_select.select_by_value('63')  # E-Mail

    # Obtener la lista de rutas de archivos adjuntos desde el archivo de Excel
    ruta_excel = excel_text.get()
    workbook = openpyxl.load_workbook(ruta_excel)
    sheet = workbook.active
    max_row = sheet.max_row

    for row in range(2, max_row + 1):
        ruta_adjunto = sheet[f'B{row}'].value
        if ruta_adjunto:
            cargar_adjunto(driver, ruta_adjunto)
            time.sleep(3)

            # Cerrar la ventana emergente después de cargar el archivo adjunto
            global ventana_adjunto
            if ventana_adjunto is not None:
                ventana_adjunto.destroy()

            # Hacer clic en el botón "Enviar"
            enviar_button = driver.find_element(
                By.XPATH, "//input[@value='Enviar']")
            enviar_button.click()

            # Obtener la fecha actual
            fecha_actual = datetime.date.today()

            # Convertir la fecha en formato de cadena
            fecha_actual_str = fecha_actual.strftime("%d/%m/%Y")

            # Esperar 3 segundos para que se complete el envío
            time.sleep(6)

            # Localizar el elemento con el ID "swal2-title" y obtener su texto
            swal2_title = driver.find_element(By.ID, "swal2-title")
            texto_radicado = swal2_title.text

            # Abrir el archivo Excel
            excel_file = excel_text.get()
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Escribir el texto en la columna C de la fila correspondiente
            fila_excel = row + 0  # Comenzar en la fila C2
            celda_excel = f"C{fila_excel}"
            sheet[celda_excel].value = f"{fecha_actual_str} {texto_radicado}"

            # Guardar el archivo Excel
            wb.save(excel_file)
            wb.close()

            # Tomar una captura de pantalla y guardarla en la ruta seleccionada para el comprobante
            ruta_comprobante = comprobante_text.get()
            nombre_excel = os.path.basename(
                excel_text.get())  # Nombre del archivo Excel
            nombre_captura = f"captura_pantalla_{row}.png"
            ruta_captura = os.path.join(ruta_comprobante, nombre_captura)
            driver.save_screenshot(ruta_captura)

            # Convertir la captura de pantalla en un archivo PDF con el nombre de la columna A
            nombre_pdf = str(sheet[f'A{row}'].value) + ".pdf"
            ruta_pdf = os.path.join(ruta_comprobante, nombre_pdf)
            convertir_imagen_a_pdf(ruta_captura, ruta_pdf)

            # Eliminar la captura de pantalla después de guardar el PDF
            os.remove(ruta_captura)

            # Esperar 10 segundos antes de volver a la página inicial y ejecutar la automatización nuevamente
            time.sleep(4)

            # Volver a la página inicial
            driver.get('https://app1.sire.gov.co/Pqrs/')

            time.sleep(3)

            # Selección del tipo de solicitante
            tipo_solicitante_select = Select(
                driver.find_element(By.ID, 'tiposoli'))
            tipo_solicitante_select.select_by_value('52')  # Perona Natural

            # Localizar el elemento de texto "adulteraciones" y hacer clic en él
            contenedor_texto = driver.find_element(By.ID, 'tratadatos')
            driver.execute_script(
                "arguments[0].scrollTop = arguments[0].scrollHeight", contenedor_texto)
            time.sleep(0.3)

            # Seleccionar la casilla de aceptación
            aceptar_checkbox = driver.find_element(By.ID, 'checkbacepto')
            aceptar_checkbox.click()

            # Hacer clic en el botón "Siguiente"
            siguiente_button = driver.find_element(By.ID, 'enviarFormIndex')
            siguiente_button.click()

            # Espera hasta que aparezca el cuadro de diálogo
            wait = WebDriverWait(driver, 10)
            modal = wait.until(EC.visibility_of_element_located(
                (By.CSS_SELECTOR, "div.modal-content")))

            # Selecciona la opción "Sí"
            opcion_si = modal.find_element(
                By.CSS_SELECTOR, "input#tratadatossi")
            driver.execute_script("arguments[0].click();", opcion_si)

            time.sleep(1)

            # Haz clic en el botón "Aceptar"
            aceptar_button = driver.find_element(
                By.XPATH, "//input[@value='Aceptar' and contains(@class, 'btn btn-primary')]")
            aceptar_button.click()

            # Selección del tipo de documento
            tipo_documento_select = Select(
                driver.find_element(By.ID, 'select_t_document'))
            tipo_documento_select.select_by_value('40')  # Tarjeta de identidad

            # Introducir el número de cédula
            numero_cedula_input = driver.find_element(By.ID, 'numero_doc')
            numero_cedula_input.send_keys('1013689567')

            # Esperar 1 segundo para que se cargue la información del usuario
            time.sleep(1)

            # Ingresar el correo de confirmación
            correo_confirmacion_input = driver.find_element(
                By.ID, 'correoconfir')
            correo_confirmacion_input.send_keys(
                'curaduria@serviciosespeciales.com.co')

            # Selección del tipo de solicitud
            tipo_solicitud_select = Select(
                driver.find_element(By.ID, 'tipo_soli'))
            tipo_solicitud_select.select_by_value('45')  # Petición

            # Selección del tema de solicitud
            tema_solicitud_select = Select(
                driver.find_element(By.ID, 'tema_soli'))
            tema_solicitud_select.select_by_value(
                '20')  # Certificado riesgo y amenaza

            # Ingresar el asunto de la solicitud
            asunto_solicitud_input = driver.find_element(By.ID, 'asunto_sol')
            asunto_solicitud_input.send_keys(
                'Se solicita al INSTITUTO DISTRITAL DE GESTION DE RIESGO Y CAMBIO CLIMATICO IDIGER, concepto en el que se indique si el predio en mención presenta afectación por inundación o deslizamiento y si esta presenta restricciones para la instalación, localización y regularización de estaciones radioeléctricas.')

            # Selección del modo de respuesta
            modo_respuesta_select = Select(
                driver.find_element(By.ID, 'modoRespuesta'))
            modo_respuesta_select.select_by_value('63')  # E-Mail

    # Cerrar el navegador
    driver.quit()


def convertir_imagen_a_pdf(ruta_imagen, ruta_pdf):
    with open(ruta_pdf, "wb") as pdf_file:
        pdf_file.write(img2pdf.convert(ruta_imagen))


# Variables para almacenar las rutas de los archivos
excel_text = tk.StringVar()
comprobante_text = tk.StringVar()
chromedriver_text = tk.StringVar()


# Campos de texto
excel_entry = tk.Entry(root, textvariable=excel_text)
excel_entry.pack(pady=5)
comprobante_entry = tk.Entry(root, textvariable=comprobante_text)
comprobante_entry.pack(pady=5)
chromedriver_entry = tk.Entry(root, textvariable=chromedriver_text)
chromedriver_entry.pack(pady=5)

# Botones
excel_button = tk.Button(root, text="Abrir Excel",
                         bg=button_bg, command=abrir_excel)
excel_button.pack(pady=5)
comprobante_button = tk.Button(
    root, text="Elegir Carpeta", bg=button_bg, command=dime_donde_guardo)
comprobante_button.pack(pady=5)
chromedriver_button = tk.Button(
    root, text="Chromedriver", bg=button_bg, command=obtener_ruta_chromedriver)
chromedriver_button.pack(pady=5)
start_button = tk.Button(
    root, text="Ejecutar Automatización", bg=button_bg, command=ejecutar_automatizacion)
start_button.pack(pady=10)

# Crear la etiqueta "Ⓡ RAMBO"
rambo_label = tk.Label(root, text="Ⓡ RAMBO", fg="orange", bg="dark green")
rambo_label.place(relx=0.85, rely=1.0, anchor="s")

# Ejecutar la interfaz gráfica
root.mainloop()
